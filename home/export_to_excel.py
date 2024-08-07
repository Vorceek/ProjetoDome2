import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from django.http import HttpResponse
from .models import TabelaPessoa
from io import BytesIO
from datetime import datetime, timedelta

def export_to_excel(request):
    # Cria um workbook e uma worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pessoas'

    # Cabeçalhos
    columns = ['Data', 'Atividade', 'Hora Início', 'Hora Fim', 'Duração']
    for col_num, column_title in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = column_title

    # Adiciona os dados
    rows = TabelaPessoa.objects.all().values_list('data', 'atividade', 'horaInicio', 'horaFim', 'duracao')
    for row_num, row in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}{row_num}']
            if isinstance(cell_value, timedelta):
                cell.value = cell_value
                cell.number_format = 'HH:MM:SS'  # Formato para duração
            elif isinstance(cell_value, (datetime.date, datetime.datetime)):
                cell.value = cell_value
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'  # Formato para data e hora
            else:
                cell.value = cell_value

    # Ajusta a largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Salva o arquivo em memória
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="tabela_pessoas.xlsx"'

    # Cria um buffer para armazenar o arquivo Excel
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())

    return response
